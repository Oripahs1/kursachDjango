import datetime

import django.http
# import reportlab.lib.pagesizes
from bs4 import BeautifulSoup
import requests

from django.views.generic import TemplateView
from django.shortcuts import render, redirect
from openpyxl.reader.excel import load_workbook

from .models import Car, PhotoCar, Worker, Order, Invoice, Duty, Price, CustomsDuty, Excise
from .forms import ParserForm, RegistrationForm, LoginForm, LogoutForm, OrderForm, OrderInOrdersForm, InvoiceForm, \
    NewInvoiceForm, DutyForm, PriceForm, CustomsDutyForm, ExciseForm
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.mixins import LoginRequiredMixin

from django.urls import reverse


class HomePageView(LoginRequiredMixin, TemplateView):
    template_name = "home.html"

    def get(self, request, *args, **kwargs):
        if request.method == 'GET':
            return render(request, self.template_name)


class LoginPageView(TemplateView):
    template_name = "registration/login.html"

    def get(self, request, *args, **kwargs):
        form = LoginForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            # user = authenticate(request, **form.cleaned_data)
            print(user)
            if user is not None:
                login(request, user)
                # messages.success(request, "Вход выполнен")
                return redirect('home')
            messages.warning(request, "Неправильное имя пользователя или пароль")
        else:
            for field in form:
                print("Field Error:", field.name, field.errors)
            messages.error(request, "Некорректная форма")
        return render(request, 'registration/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('home')


class RegistrationPageView(TemplateView):
    template_name = "registration/registration.html"

    def get(self, request, *args, **kwargs):
        if request.method == 'GET':
            form = RegistrationForm()
            form.fields['date_joined'].widget.attrs.update({'value': datetime.date.today()})
            form.fields['is_active'].widget.attrs.update({'value': True})
            form.fields['is_staff'].widget.attrs.update({'value': True})
            form.fields['is_superuser'].widget.attrs.update({'value': True})
            return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            form = RegistrationForm(request.POST)
            if form.is_valid():
                if form.username_clean() is None:
                    messages.warning(request, "Данное имя пользователя уже используется")
                    return render(request, 'registration/registration.html', {'form': form})
                if form.passport_clean() is None:
                    messages.warning(request, "Пользователь с таким паспортом уже существует")
                    return render(request, 'registration/registration.html', {'form': form})
                if form.clean_password2() is None:
                    messages.warning(request, "Пароли не совпадают")
                    return render(request, 'registration/registration.html', {'form': form})
                if form.save() == None:
                    messages.warning(request,
                                     "Этот номер паспорта уже используется. Пожалуйста, выберите другой номер паспорта.")
                    return render(request, 'registration/registration.html', {'form': form})
                else:
                    form.save()
                    messages.success(request, "Пользователь зарегистрирован")
            else:
                messages.error(request, "Некорректная форма")
        else:
            form = RegistrationForm()
        return render(request, 'registration/registration.html', {'form': form})


class CustomsDutysPageView(TemplateView):
    template_name = 'customs_dutys.html'

    def get(self, request, *args, **kwargs):
        customs_dutys = CustomsDuty.objects.all()
        return render(request, self.template_name, {'customs_dutys': customs_dutys})


class CustomsDutyNewPageView(TemplateView):
    template_name = 'customs_duty.html'

    def get(self, request, *args, **kwargs):
        print('NEW')
        form = CustomsDutyForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            form = CustomsDutyForm(request.POST)
            if form.is_valid():
                form.save()
                messages.info(request, "Добавлена новая пошлина")
            else:
                for field in form:
                    print("Field Error:", field.name, field.errors)
                messages.info(request, "Ошибка валидации формы")
        customs_dutys = CustomsDuty.objects.all()
        return render(request, 'customs_dutys.html', {'customs_dutys': customs_dutys})


class CustomsDutyPageView(TemplateView):
    template_name = 'customs_duty.html'

    def get(self, request, *args, **kwargs):
        form = CustomsDutyForm()
        # form.fields['name'].initial
        print('Просим')
        customs_duty = CustomsDuty.objects.get(id=kwargs['customs_duty_id'])
        form.fields['type'].initial = customs_duty.type
        form.fields['value_first'].initial = customs_duty.value_first
        form.fields['value_last'].initial = customs_duty.value_last
        form.fields['bet'].initial = customs_duty.bet
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            form = CustomsDutyForm(request.POST)
            if form.is_valid():
                print(kwargs['customs_duty_id'])

                form.update(kwargs['customs_duty_id'])
                messages.info(request, "Добавлена новая пошлина")
            else:
                for field in form:
                    print("Field Error:", field.name, field.errors)
                messages.info(request, "Ошибка валидации формы")
        customs_dutys = CustomsDuty.objects.all()
        return render(request, 'customs_dutys.html', {'customs_dutys': customs_dutys})


class ExcisesPageView(TemplateView):
    template_name = 'excises.html'

    def get(self, request, *args, **kwargs):
        excises = Excise.objects.all()
        return render(request, self.template_name, {'excises': excises})


class ExciseNewPageView(TemplateView):
    template_name = 'excise.html'

    def get(self, request, *args, **kwargs):
        print('NEW')
        form = ExciseForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            form = ExciseForm(request.POST)
            if form.is_valid():
                form.save()
                messages.info(request, "Добавлена новая пошлина")
            else:
                for field in form:
                    print("Field Error:", field.name, field.errors)
                messages.info(request, "Ошибка валидации формы")
        excises = Excise.objects.all()
        return render(request, 'excises.html', {'excises': excises})


class ExcisePageView(TemplateView):
    template_name = 'excise.html'

    def get(self, request, *args, **kwargs):
        form = ExciseForm()
        # form.fields['name'].initial
        excise = Excise.objects.get(id=kwargs['excise_id'])
        form.fields['power_first_car'].initial = excise.power_first_car
        form.fields['power_last_car'].initial = excise.power_last_car
        form.fields['bet'].initial = excise.bet
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            form = ExciseForm(request.POST)
            if form.is_valid():
                print(kwargs['excise_id'])

                form.update(kwargs['excise_id'])
                messages.info(request, "Добавлена новая пошлина")
            else:
                for field in form:
                    print("Field Error:", field.name, field.errors)
                messages.info(request, "Ошибка валидации формы")
        excises = Excise.objects.all()
        return render(request, 'excises.html', {'excises': excises})


class PricesPageView(TemplateView):
    template_name = 'prices.html'

    def get(self, request, *args, **kwargs):
        prices = Price.objects.all()
        return render(request, self.template_name, {'prices': prices})


class PriceNewPageView(TemplateView):
    template_name = 'price.html'

    def get(self, request, *args, **kwargs):
        print('NEW')
        form = PriceForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            form = PriceForm(request.POST)
            if form.is_valid():
                form.save()
                messages.info(request, "Добавлена новая пошлина")
            else:
                for field in form:
                    print("Field Error:", field.name, field.errors)
                messages.info(request, "Ошибка валидации формы")
        prices = Price.objects.all()
        return render(request, 'prices.html', {'prices': prices})


class PricePageView(TemplateView):
    template_name = 'price.html'

    def get(self, request, *args, **kwargs):
        form = PriceForm()
        # form.fields['name'].initial
        price = Price.objects.get(id=kwargs['price_id'])
        form.fields['price_first_car'].initial = price.price_first_car
        if price.price_last_car is not None:
            form.fields['price_last_car'].initial = price.price_last_car
        # else:
        #     form.fields['']
        form.fields['price_transportation'].initial = price.price_transportation
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            form = PriceForm(request.POST)
            if form.is_valid():
                print(kwargs['price_id'])

                form.update(kwargs['price_id'])
                messages.info(request, "Добавлена новая пошлина")
            else:
                for field in form:
                    print("Field Error:", field.name, field.errors)
                messages.info(request, "Ошибка валидации формы")
        prices = Price.objects.all()
        return render(request, 'prices.html', {'prices': prices})


class DutiesPageView(TemplateView):
    template_name = 'duties.html'

    def get(self, request, *args, **kwargs):
        duties = Duty.objects.all()
        return render(request, self.template_name, {'duties': duties})


class DutyNewPageView(TemplateView):
    template_name = 'duty.html'

    def get(self, request, *args, **kwargs):
        form = DutyForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            form = DutyForm(request.POST)
            if form.is_valid():
                form.save()
                messages.info(request, "Добавлена новая пошлина")
            else:
                for field in form:
                    print("Field Error:", field.name, field.errors)
                messages.info(request, "Ошибка валидации формы")
        duties = Duty.objects.all()
        return render(request, 'duties.html', {'duties': duties})


class DutyPageView(TemplateView):
    template_name = 'duty.html'

    def get(self, request, *args, **kwargs):
        form = DutyForm()
        # form.fields['name'].initial
        duty = Duty.objects.get(id=kwargs['duty_id'])
        form.fields['volume_first'].initial = duty.volume_first
        form.fields['volume_last'].initial = duty.volume_last
        form.fields['coefficient_less_3'].initial = duty.coefficient_less_3
        form.fields['coefficient_more_3'].initial = duty.coefficient_more_3
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            form = DutyForm(request.POST)
            if form.is_valid():
                print(kwargs['duty_id'])

                form.update(kwargs['duty_id'])
                messages.info(request, "Добавлена новая пошлина")
            else:
                for field in form:
                    print("Field Error:", field.name, field.errors)
                messages.info(request, "Ошибка валидации формы")
        duties = Duty.objects.all()
        return render(request, 'duties.html', {'duties': duties})


class WorkersPageView(TemplateView):
    template_name = "workers.html"

    def get(self, request, *args, **kwargs):
        workers = Worker.objects.filter(is_superuser=False)
        return render(request, 'workers.html', {'workers': workers})


class WorkersCardPageView(TemplateView):
    template_name = "worker_card.html"

    def get(self, request, *args, **kwargs):
        worker = Worker.objects.get(id=kwargs.get('worker_id'))
        worker_data = Worker.objects.all()
        form = RegistrationForm()
        form.fields['username'].widget.attrs.update({'value': worker.username})
        form.fields['full_name'].widget.attrs.update({'value': worker.full_name})
        # Вот тут хуй знает как сделать не нашел
        form.fields['job_title'].widget.attrs.update({'value': worker.job_title})
        # Вот тут хуй знает как сделать не нашел
        form.fields['passport'].widget.attrs.update({'value': worker.passport})
        form.fields['phone_number'].widget.attrs.update({'value': worker.phone_number})
        form.fields['password'].widget.attrs.update({'value': worker.password})
        form.fields['password2'].widget.attrs.update({'value': worker.password})

        return render(request, 'worker_card.html', {'worker': worker, 'worker_data': worker_data, 'form': form})

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            form = RegistrationForm(request.POST)
            if form.is_valid():
                # form.username_clean()
                if form.clean_password2() is None:
                    # message = messages.info(request, 'Your password has been changed successfully!')
                    messages.warning(request, "Пароли не совпадают")
                    return render(request, 'registration/registration.html', {'form': form})
                if form.passport_clean() is None:
                    # message = messages.info(request, 'Your password has been changed successfully!')
                    messages.warning(request, "Пользователь с таким паспортом уже существует")
                    return render(request, 'registration/registration.html', {'form': form})
                if form.username_clean() is None:
                    # message = messages.info(request, 'Your password has been changed successfully!')
                    messages.warning(request, "Пользователь с таким именем уже существует")
                    return render(request, 'registration/registration.html', {'form': form})
                form.update()
                messages.success(request, "Данные обновлены")
            else:
                messages.error(request, "Некорректная форма")

        else:
            form = RegistrationForm()
        return render(request, 'registration/registration.html', {'form': RegistrationForm()})


class OrderInOrdersPageView(TemplateView):
    template_name = 'order_in_orders.html'

    def get(self, request, *args, **kwargs):
        order = Order.objects.get(id_order=kwargs['order_id'])
        form = OrderInOrdersForm()

        form.fields['id_order'].widget.attrs.update({'value': order.id_order})
        form.fields['first_name_client'].widget.attrs.update({'value': order.id_customer.first_name_client})
        form.fields['last_name_client'].widget.attrs.update({'value': order.id_customer.last_name_client})
        form.fields['patronymic_client'].widget.attrs.update({'value': order.id_customer.patronymic_client})
        form.fields['telephone'].widget.attrs.update({'value': order.id_customer.telephone})
        form.fields['date_start'].widget.attrs.update({'value': order.date_start})
        if order.price is not None:
            form.fields['price'].widget.attrs.update({'value': str(order.price) + ' р.'})
        form.fields['sbts'].widget.initial_text = ''
        form.fields['sbts'].widget.input_text = 'Заменить'
        form.fields['ptd'].widget.initial_text = ''
        form.fields['ptd'].widget.input_text = 'Заменить'
        form.fields['ptd'].widget.clear_checkbox_label = ''
        form.fields['sbts'].widget.clear_checkbox_label = ''
        if order.date_end is not None:
            form.fields['date_end'].widget.attrs.update({'value': order.date_end, 'readonly': 'True'})
        if order.comment is not None:
            form.fields['comment'].initial = order.comment
        if order.sbts is not None:
            form.fields['sbts'].initial = order.sbts
        if order.ptd is not None:
            form.fields['ptd'].initial = order.ptd
        return render(request, self.template_name, {'order': order, 'form': form, 'order_id': order.id_order})

    def post(self, request, *args, **kwargs):
        print(request.FILES)
        if request.method == 'POST' and 'update' in request.POST:
            form = OrderInOrdersForm(request.POST, request.FILES)
            if form.is_valid():

                order = Order.objects.filter(id_order=form.cleaned_data['id_order'])
                order = order[0]
                # is_initial
                print(form.initial)
                print(form.cleaned_data['date_end'])
                # print(form.fields['ptd'].initial)

                if order.ptd == '':
                    order.ptd = request.FILES.get('ptd')
                else:
                    order.ptd = order.ptd

                if order.sbts == '':
                    order.sbts = request.FILES.get('sbts')
                else:
                    order.sbts = order.sbts

                order.save()

                messages.success(request, "Заказ изменен")
                form.save()
            else:
                messages.error(request, "Некорректная форма")
                for field in form:
                    print("Field Error:", field.name, field.errors)

        elif request.method == 'POST' and 'calculate_price' in request.POST:
            form = OrderInOrdersForm(request.POST, request.FILES)
            print('считаем цену')
            order = Order

            if form.is_valid():
                order = Order.objects.get(id_order=form.cleaned_data['id_order'])
                price = form.cleaned_data['price_for_buhgalter']
                power = int(form.cleaned_data['power'])
                duties = Duty.objects.all()
                prices = Price.objects.all()
                excises = Excise.objects.all()
                customs_dutys = CustomsDuty.objects.all()
                volume = int(order.id_car.volume) / 1000
                year = int(datetime.datetime.now().year) - int(order.id_car.year_car)
                base_bet = 20000
                coefficient_bet = 0
                coefficient_excise = 0
                coefficient_customs_duty = 0
                price_transportation = 0
                for duty in duties:
                    if duty.volume_first <= volume <= duty.volume_last and year > 3:
                        coefficient_bet = duty.coefficient_more_3
                    elif duty.volume_first <= volume <= duty.volume_last and year <= 3:
                        coefficient_bet = duty.coefficient_less_3
                #
                for price_data in prices:
                    if price_data.price_first_car <= int(price) < price_data.price_last_car:
                        price_transportation = price_data.price_transportation
                    elif price_data.price_last_car == 0 and price_transportation == 0:
                        price_transportation = price_data.price_transportation
                volume_or_price = 0
                if year < 3:
                    customs_dutys = CustomsDuty.objects.filter(type=CustomsDuty.TYPE_CHOICE[0][1])
                    volume_or_price = price
                elif 3 <= year < 5:
                    customs_dutys = CustomsDuty.objects.filter(type=CustomsDuty.TYPE_CHOICE[1][1])
                    volume_or_price = volume * 1000
                else:
                    customs_dutys = CustomsDuty.objects.filter(type=CustomsDuty.TYPE_CHOICE[2][1])
                    volume_or_price = volume * 1000

                for customs_duty in customs_dutys:
                    if customs_duty.value_first <= int(volume_or_price) <= customs_duty.value_last:
                        coefficient_customs_duty = int(volume) * 1000 * customs_duty.bet
                        print(volume * 1000, year, customs_duty.bet)
                    elif customs_duty.value_last == 0 and coefficient_customs_duty == 0:
                        coefficient_customs_duty = int(volume) * 1000 * customs_duty.bet

                for excise in excises:
                    if excise.power_first_car <= power <= excise.power_last_car:
                        coefficient_excise = power * excise.bet
                        print(coefficient_excise)
                    elif coefficient_excise == 0 and excise.power_last_car == 0:
                        coefficient_excise = power * excise.bet

                # Ставки утилизационного сбора
                final_price = int(price) + base_bet * coefficient_bet
                # Цена доставки тачки из Японии
                final_price = final_price + price_transportation
                # Таможенная пошлина customs_dutys
                final_price = final_price + coefficient_customs_duty
                # Акциз excises
                final_price = final_price + coefficient_excise
                # НДС (стоимость авто+таможенная пошлина+акциз)*20%
                nds = (int(price) + int(coefficient_customs_duty) + int(coefficient_excise)) * 0.2
                final_price = final_price + int(nds)
                print(power)
                print(int(price), base_bet * coefficient_bet, price_transportation, coefficient_excise,
                      coefficient_customs_duty, nds)
                print(final_price)
                form.fields['price'].widget.attrs.update({'value': final_price})

            return render(request, 'order_in_orders.html', {'form': form, 'order': order})

        user_id = request.user.id
        orders = Order.objects.filter(date_end=None, id_worker=user_id)
        return render(request, 'orders.html', {"orders": orders})


class OrdersPageView(TemplateView):
    template_name = "orders.html"

    def get(self, request, *args, **kwargs):
        user_id = request.user.id
        orders = Order.objects.filter(date_end=None, id_worker=user_id)
        return render(request, 'orders.html', {'orders': orders})


class OrderPageView(TemplateView):
    template_name = "order.html"

    def get(self, request, *args, **kwargs):
        car = Car.objects.get(id_car=kwargs.get('car_id'))
        form = OrderForm()
        photo = PhotoCar.objects.filter(id_car=kwargs.get('car_id'))[:1][0].photo
        form.fields['id_car'].widget.attrs.update({'value': car.id_car})
        user_name = Worker.objects.filter(id=request.user.id)[0]
        form.fields['worker'].widget.attrs.update({'value': user_name})
        form.fields['price'].initial = str(car.price) + ' р.'
        return render(request, 'order.html', {'car': car, 'form': form, 'photo': photo})

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            form = OrderForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Заказ создан")
                return django.http.HttpResponseRedirect(reverse('orders'))
            else:
                messages.error(request, "Некорректная форма")
                return render(request, 'order.html', {'form': form})
        else:
            form = OrderForm()
        user_id = request.user.id
        orders = Order.objects.filter(date_end=None, id_worker=user_id)
        return render(request, 'orders.html', {'orders': orders})


class CatalogPageView(TemplateView):

    def get(self, request, *args, **kwargs):
        cars = Car.objects.all()
        cars = Car.objects.filter(auc_date__gte=datetime.date.today())

        for el in cars:
            el.image = PhotoCar.objects.filter(id_car=el.id_car)[:1][0].photo

        return render(request, 'catalog.html', {'cars': cars})


class CarPageView(TemplateView):
    template_name = "car.html"

    def get(self, request, *args, **kwargs):
        car = Car.objects.get(id_car=kwargs.get('car_id'))
        photo = PhotoCar.objects.filter(id_car=car)

        # Достаем данные из excel
        file_path = 'cars_price.xlsx'
        workbook = load_workbook(filename=file_path)

        models_list = workbook.worksheets[1]
        mark_list = workbook.worksheets[0]

        models = []
        skip_first_row = False
        for row in models_list.iter_rows(values_only=True, min_row=2 if skip_first_row else 1):
            model = {
                'id': row[0],
                'mark': row[1],
                'model': row[2],
                'price': row[3],
            }
            models.append(model)
        marks = []
        for row in mark_list.iter_rows(values_only=True, min_row=2 if skip_first_row else 1):
            mark = {
                'id': row[0],
                'mark': row[1]
            }
            marks.append(mark)
        # Достаем данные из excel

        # Узнаем цену машины
        car_for_test = Car.objects.all()
        for car_test in car_for_test:
            car_title = car_test.title.split()
            is_car = False
            price = 0
            for car_for_test in models:

                # if str(marks[int(car_for_test['mark'])-1]['mark']) == 'Acura':
                #     print(car_for_test, marks[int(car_for_test['mark'])-1]['mark'], 'true', int(car_for_test['mark']))
                # print(marks[int(car_for_test['mark'])-1]['mark'], car_title[0].lower())
                if (str(car_for_test['model']).lower() == car_title[1].lower()
                        and
                        str(marks[int(car_for_test['mark']) - 1]['mark']).lower() == car_title[0].lower()):
                    is_car = True
                    price = car_for_test['price']
            if is_car is False:
                price = 'Нет информации о цене'

            print(car_test.title, price)

        return render(request, 'car.html', {'car': car, 'photo': photo})


class ParserPageView(TemplateView):
    template_name = "parser.html"

    def get(self, request, *args, **kwargs):
        form = ParserForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        linked_list = list()
        print('Пришел пост запрос')
        if request.method == 'POST' and 'import' in request.POST:
            url = 'https://www.carwin.ru/japanauc/'
            response = requests.get(url)
            html_page = BeautifulSoup(response.text, 'lxml')
            urls_list = html_page.find('ul', 'pagination')
            urls_list = urls_list.find_all('a')

            for i in range(1, len(urls_list) + 1):

                response = requests.get(url + str(i))
                html_page = BeautifulSoup(response.text, 'lxml')
                links = html_page.find_all('a', 'pic')
                for link in links:
                    linked_list.append(link['href'])

            url = 'https://www.carwin.ru'
            print(url + linked_list[0])
            k = 0
            for link in linked_list:
                k += 1
                print(k)
                self.link_obr(url + link)
            messages.success(request, "База данных обновлена")

        return render(request, 'parser.html', {'response': 'success'})

    def link_obr(self, url):

        response = requests.get(url)
        html_page = BeautifulSoup(response.text, "lxml")
        number_of_auc = html_page.find('div', 'row_desc_middle')

        car = Car.objects.filter(auc_number=number_of_auc.text)

        if html_page.find('div', 'page_title') is not None and not car:

            Obj = Car_data(url)
            print(url)
            Obj.print()

            Obj.save_me_to_bd()
            del Obj
        else:

            print('Было, знаем!', url)
            # return


class Parser(object):
    html_page = None

    def __init__(self, url):
        response = requests.get(url)
        self.html_page = BeautifulSoup(response.text, "lxml")
        print(response)

    def parse_title(self):
        name = self.html_page.find('div', 'page_title').text
        # Достаем данные из excel
        file_path = 'cars_price.xlsx'
        workbook = load_workbook(filename=file_path)

        models_list = workbook.worksheets[1]
        mark_list = workbook.worksheets[0]

        models = []
        skip_first_row = False
        for row in models_list.iter_rows(values_only=True, min_row=2 if skip_first_row else 1):
            model = {
                'id': row[0],
                'mark': row[1],
                'model': row[2],
                'price': row[3],
            }
            models.append(model)
        marks = []
        for row in mark_list.iter_rows(values_only=True, min_row=2 if skip_first_row else 1):
            mark = {
                'id': row[0],
                'mark': row[1]
            }
            marks.append(mark)

        # Достаем данные из excel

        # Узнаем цену машины

        car_title = name.split()
        is_car = False
        price = 0
        for car_for_test in models:

            # if str(marks[int(car_for_test['mark'])-1]['mark']) == 'Acura':
            #     print(car_for_test, marks[int(car_for_test['mark'])-1]['mark'], 'true', int(car_for_test['mark']))
            # print(marks[int(car_for_test['mark'])-1]['mark'], car_title[0].lower())
            if (str(car_for_test['model']).lower() == car_title[1].lower()
                    and
                    str(marks[int(car_for_test['mark']) - 1]['mark']).lower() == car_title[0].lower()):
                is_car = True
                price = car_for_test['price']
        if is_car is False:
            price = 'Нет информации о цене'

        print(name, price, '______________________________________________________')
        price_name = {'name': name, 'price': price}
        return price_name

    def parse_auction_data(self):
        auction_data = self.html_page.find('div', 'col_left').text
        auction_data = [value for value in auction_data.split('\n') if value != '']
        return auction_data

    def parse_car_options(self):
        car_options = self.html_page.find('div', 'car_description')
        car_options = car_options.find_all('div', 'car_option')

        data_set = {'year_car': '', 'mileage': '', 'color': '', 'options': '', 'the_body': '', 'volume': '', 'cpp': '',
                    'estimation': ''}

        # Условие не трогать, работает и слава богу
        for el in range(0, len(car_options)):
            split_data_of_car = car_options[el].text.split()

            match split_data_of_car[0]:
                case 'Год':
                    if len(split_data_of_car) > 1:
                        data_set['year_car'] = split_data_of_car[1]
                case 'Пробег':
                    if len(split_data_of_car) > 1:
                        data_set['mileage'] = split_data_of_car[1]
                case 'Цвет':
                    if len(split_data_of_car) > 1:
                        data_set['color'] = split_data_of_car[1]
                case 'Опции':
                    if len(split_data_of_car) > 1:
                        data_set['options'] = split_data_of_car[1]
                case 'Кузов':
                    if len(split_data_of_car) > 1:
                        data_set['the_body'] = split_data_of_car[1]
                case 'Объем':
                    if len(split_data_of_car) > 1:
                        data_set['volume'] = split_data_of_car[1]
                case 'КПП':
                    if len(split_data_of_car) > 1:
                        data_set['cpp'] = split_data_of_car[1]
                case 'Оценка':
                    if len(split_data_of_car) > 1:
                        data_set['estimation'] = split_data_of_car[1]

        car_options = data_set
        return car_options

    def parse_content(self):
        content = self.html_page.find('div', 'content')
        content = content.find_all('td')

        data_set = {'cooling': '',
                    'set': '',
                    'result': '',
                    'start_price': '',
                    'transmission': '',
                    'location_auction': '',
                    'year': '',
                    'alt_color': '',
                    'condition': '',
                    'fuel': '',
                    'equipment': '',
                    'deadline_for_the_price_offer': '',
                    'day_of_the_event': '',
                    'number_of_sessions': ''}
        # Почему здесь по-другому, не знаю, но тоже работает и слава богу
        for el in range(0, len(content), 2):
            match content[el].text:
                case ' охлаждение ':
                    data_set['cooling'] = content[el + 1].text
                case ' комплектация ':
                    data_set['set'] = content[el + 1].text
                case ' результат ':
                    data_set['result'] = content[el + 1].text
                case ' старт ':
                    data_set['start_price'] = content[el + 1].text
                case ' коробка передач ':
                    data_set['transmission'] = content[el + 1].text
                case ' место проведения ':
                    data_set['location_auction'] = content[el + 1].text
                case ' год ':
                    data_set['year'] = content[el + 1].text
                case ' цвет ':
                    data_set['alt_color'] = content[el + 1].text
                case ' состояние ':
                    data_set['condition'] = content[el + 1].text
                case ' топливо ':
                    data_set['fuel'] = content[el + 1].text
                case ' оборудование ':
                    data_set['equipment'] = content[el + 1].text
                case ' конечный срок предложения цены ':
                    data_set['deadline_for_the_price_offer'] = content[el + 1].text
                case ' день проведения ':
                    data_set['day_of_the_event'] = content[el + 1].text
                case ' количество проведений ':
                    data_set['number_of_sessions'] = content[el + 1].text

        content = data_set
        return content

    def parse_image(self):
        image = self.html_page.find('div', 'my-gallery')
        image = image.find_all('img')
        form_data = list()

        for el in range(len(image)):
            form_data.append(image[el]['src'])
        image = form_data
        return image

    def parse_auc_list(self):
        auc_list = self.html_page.find('div', 'scheme_block')
        auc_list = auc_list.find('img')
        auc_list = auc_list['src']
        return auc_list


class Car_data(object):
    title = ''
    auction_data = ''
    car_options = ''
    content = ''
    auc_link = ''
    image = ''
    # auction_data
    auc_name = ''
    auc_number = ''
    auc_date = ''
    # car_options
    year_car = ''
    mileage = ''
    color = ''
    options = ''
    the_body = ''
    volume = ''
    cpp = ''
    estimation = ''
    # content
    cooling = ''
    set = ''
    result = ''
    start_price = ''
    transmission = ''
    location_auction = ''
    year = ''
    alt_color = ''
    condition = ''
    fuel = ''
    equipment = ''
    deadline_for_the_price_offer = ''
    day_of_the_event = ''
    number_of_sessions = ''

    auc_list = ''

    price = ''

    def __init__(self, url):
        parser = Parser(url)
        self.auc_link = url
        # для Car_of_page
        price_name = parser.parse_title()
        self.title = price_name['name']
        self.price = price_name['price']
        self.auction_data = parser.parse_auction_data()
        self.car_options = parser.parse_car_options()
        self.content = parser.parse_content()
        self.image = parser.parse_image()

        # для Car_data

        # auction_data
        self.auc_name = parser.parse_auction_data()[0]
        self.auc_number = parser.parse_auction_data()[1]
        self.auc_date = parser.parse_auction_data()[2]

        # car_options
        form_data = parser.parse_car_options()

        self.year_car = form_data['year_car']
        self.mileage = form_data['mileage']
        self.color = form_data['color']
        self.options = form_data['options']
        self.the_body = form_data['the_body']
        self.volume = form_data['volume']
        self.cpp = form_data['cpp']
        self.estimation = form_data['estimation']

        # content
        form_data = parser.parse_content()

        self.cooling = form_data['cooling']
        self.set = form_data['set']
        self.result = form_data['result']
        self.start_price = form_data['start_price']
        self.transmission = form_data['transmission']
        self.location_auction = form_data['location_auction']
        self.year = form_data['year']
        self.alt_color = form_data['alt_color']
        self.condition = form_data['condition']
        self.fuel = form_data['fuel']
        self.equipment = form_data['equipment']
        self.deadline_for_the_price_offer = form_data['deadline_for_the_price_offer']
        self.day_of_the_event = form_data['day_of_the_event']
        self.number_of_sessions = form_data['number_of_sessions']

        self.auc_list = parser.parse_auc_list()

    def print(self):
        print('название машины', self.title, 'аукцион', self.auction_data, 'основное про машину', self.car_options,
              'таблица', self.content, sep='\n', end='\n')
        print('картинки', self.image, end='\n')
        print(self.auc_name, self.auc_number, self.auc_date, sep='\n', end='\n')
        print(self.year_car, self.mileage, self.color, self.options, self.the_body, self.volume, self.cpp,
              self.estimation, sep='\n', end='\n')
        print(self.cooling, self.condition, self.fuel, self.equipment)

    def __del__(self):
        print('Удален')

    def save_me_to_bd(self):
        new_car_new = Car.objects.create(
            auc_link=self.auc_link,
            title=self.title,
            auc_name=self.auc_name,
            auc_number=self.auc_number,
            auc_date=self.auc_date,
            year_car=self.year_car,
            mileage=self.mileage,
            color=self.color,
            options=self.options,
            the_body=self.the_body,
            volume=self.volume,
            cpp=self.cpp,
            estimation=self.estimation,
            cooling=self.cooling,
            set=self.set,
            result=self.result,
            start_price=self.start_price,
            transmission=self.transmission,
            location_auction=self.location_auction,
            year=self.year,
            alt_color=self.alt_color,
            condition=self.condition,
            fuel=self.fuel,
            equipment=self.equipment,
            deadline_for_the_price_offer=self.deadline_for_the_price_offer,
            day_of_the_event=self.day_of_the_event,
            number_of_sessions=self.number_of_sessions,
            auc_list=self.auc_list,
            price=self.price
        )
        for el in range(len(self.image)):
            PhotoCar.objects.create(id_car=new_car_new, photo=self.image[el])

        print(new_car_new)


class BuhgalterPageView(TemplateView):
    template_name = "buhgalter/buhgalter.html"

    def get(self, request, *args, **kwargs):
        invoices = Invoice.objects.all()
        print('Тут должен быть инвойс')
        print(invoices)
        return render(request, 'buhgalter/buhgalter.html', {'invoices': invoices})


class BuhgalterInvoicePageView(TemplateView):
    template_name = "buhgalter/invoice.html"

    def get(self, request, *args, **kwargs):
        invoice = Invoice.objects.get(id_invoice=kwargs.get('invoice_id'))
        invoice_data = Invoice.objects.all()
        form = InvoiceForm()
        form.fields['id_invoice'].widget.attrs.update({'value': invoice.id_invoice})
        form.fields['payer'].widget.attrs.update({'value': invoice.payer})
        form.fields['seller'].widget.attrs.update({'value': invoice.seller})
        form.fields['date_form'].widget.attrs.update({'value': invoice.date_form})
        form.fields['date_pay'].widget.attrs.update({'value': invoice.date_pay})
        form.fields['sum'].widget.attrs.update({'value': invoice.sum})
        form.fields['check_document'].widget.attrs.update({'value': invoice.check_document})
        form.fields['assigning'].widget.attrs.update({'value': invoice.assigning})
        form.fields['scan'].widget.attrs.update({'value': invoice.scan})
        form.fields['type'].widget.attrs.update({'value': invoice.type})

        return render(request, 'buhgalter/invoice.html',
                      {'invoice': invoice, 'invoice_data': invoice_data, 'form': form})

    def post(self, request, *args, **kwargs):
        invoices = Invoice.objects.all()
        if request.method == 'POST':
            form = InvoiceForm(request.POST)
            print('Валидная или инвалидная форма', form.is_valid())
            print(form.errors)
            if form.is_valid():
                form.update()
                messages.success(request, "Данные обновлены")
            else:
                messages.error(request, "Некорректная форма")
        else:
            form = InvoiceForm()
        return render(request, 'buhgalter/buhgalter.html', {'invoices': invoices})


class BuhgalterNewInvoicePageView(TemplateView):
    template_name = "buhgalter/new_invoice.html"

    def get(self, request, *args, **kwargs):
        if request.method == 'GET':
            form = InvoiceForm()
            return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):
        invoices = Invoice.objects.all()
        if request.method == 'POST':
            print('Запрос пришел')
            form = NewInvoiceForm(request.POST)
            print(form.is_valid())
            print(form.errors)
            if form.is_valid():
                form.save()
                messages.success(request, "Счет на оплату сохранен")
                return render(request, 'buhgalter/new_invoice.html', {'form': form})
            else:
                messages.error(request, "Некорректная форма")
            # form.save()
            # messages.info(request, "Счет на оплату сохранен")
        else:
            form = NewInvoiceForm()
        return render(request, 'buhgalter/buhgalter.html', {'invoices': invoices})


from django.http import HttpResponse
from PyPDF2 import PdfFileReader, PdfFileWriter, PdfFileMerger
import io
from reportlab.pdfgen import canvas
from django.http import FileResponse
from reportlab.lib.colors import black
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import inch, cm
from reportlab.lib.pagesizes import letter


def draw_multiline_text(page, text, x, y, width, line_height, font_name, font_size, align='left'):
    """
    Рисует многострочный текст на странице PDF с заданным выравниванием
    """
    page.setFont(font_name, font_size)
    lines = []
    current_line = ''
    for word in text.split():
        # if "ПЕРЕХОД" in word:
        #     continue
        if page.stringWidth(current_line + ' ' + word) < width and word != "ПЕРЕХОД":
            current_line += ' ' + word if current_line else word
        else:
            lines.append(current_line)
            if word == "ПЕРЕХОД":
                current_line = ''
                continue
            current_line = word
    lines.append(current_line)

    total_height = len(lines) * line_height
    if align == 'center':
        y += (line_height - total_height) / 2
    elif align == 'right':
        y += line_height - total_height

    for line in lines:
        page.drawString(x, y, line.strip())
        y -= line_height


def some_view(request, order_id):
    buffer = io.BytesIO()
    page = canvas.Canvas(buffer, pagesize=letter)
    order = Order.objects.get(id_order=order_id)

    pdfmetrics.registerFont(TTFont('Arial', 'static\\fonts\\ArialRegular.ttf'))
    pdfmetrics.registerFont(TTFont('Arial-Bold', 'static\\fonts\\ArialBold.ttf'))
    page.setFont("Arial", 12)
    page.drawString(9 * cm, 26.5 * cm, f"Агентский договор № {order.id_order} /24")
    page.setFont("Arial", 8)
    page.drawString(6 * cm, 26.5 * cm - 0.4 * cm,
                    f"на приобретение транспортного средства, его доставку в РФ и оформление")
    page.setFont("Arial", 12)
    page.drawString(2.5 * cm, 26.5 * cm - cm,
                    f"г. Владивосток                                                                                                   {order.date_start}")
    text = (
        f"Общество с ограниченной ответственностью «NAHODKA MOTORS»,"
    )
    draw_multiline_text(page, text, 4 * cm, 26.5 * cm - 0.5 * cm * 3, 17 * cm, 0.5 * cm, "Arial", 12)
    text = (
        f"именуемое в тексте договора «Поставщик», в лице ____________________, действующего на основании {order.id_customer} с одной стороны,"
        f"и ______________________, дата рождения {order.id_customer.date_of_birth} г, паспорт "
        f"{order.id_customer.passport_number} №{order.id_customer.passport_series}, выдан {order.id_customer.passport_department_name},"
        f"код подразделения {order.id_customer.passport_department_code},именуемый в "
        f"тексте договора «Заказчик», с другой стороны, заключили настоящий договор о нижеследующем:"
    )
    draw_multiline_text(page, text, 2.5 * cm, 26.5 * cm - 0.5 * cm * 4, 17 * cm, 0.5 * cm, "Arial", 12)
    text = (
        f"1. Предмет договора"
    )
    draw_multiline_text(page, text, 2.5 * cm, 26.5 * cm - 0.5 * cm * 9, 17 * cm, 0.5 * cm, "Arial-Bold", 12)
    page.setFont("Arial", 12)
    text = (
        "тексте договора «Заказчик», с другой стороны, заключили настоящий договор о нижеследующем:"
        " ПЕРЕХОД 1.1. Поставщик обязуется за вознаграждение совершать по поручению Заказчика юридические и иные действия от "
        "своего имени, но за счет Заказчика, либо от имени и за счет Заказчика."
        " ПЕРЕХОД 1.2. Поставщик приобретает права и становится обязанным по сделке, совершенной с третьим лицом от своего имени "
        "за счет Заказчика."
        " ПЕРЕХОД 1.3. По сделке, совершенной Поставщиком с третьим лицом от имени и за счет Заказчика, права и обязанности "
        "возникают у Заказчика."
        " ПЕРЕХОД 1.4. В соответствии с настоящим договором Поставщик обязуется по поручению Заказчика организовать покупку "
        "транспортного средства (далее по тексту ТС) на автомобильных аукционах в Японии и доставку указанного ТС до "
        "места получения ТС в соответствии с заявкой (поручением) Заказчика."
        " ПЕРЕХОД 1.5. Для исполнения поручения Заказчика Поставщик обязуется совершить следующие действия:"
        " ПЕРЕХОД - осуществить покупку указанного Заказчиком ТС на аукционе Японии;"
        " ПЕРЕХОД - осуществить доставку приобретенного ТС в порт погрузки в Японии;"
        " ПЕРЕХОД - осуществить доставку приобретенного ТС морским транспортом до порта г. Владивосток;"
        " ПЕРЕХОД - осуществить действия по таможенной очистке ТС в г. Владивосток, в том числе оформить необходимые таможенные "
        "документы;"
        " ПЕРЕХОД - осуществить передачу приобретенного ТС Заказчику."
        " ПЕРЕХОД 1.6. Для осуществления действий указанных в п.1.5. настоящего договора Поставщик заключает от своего имени "
        "необходимые договоры, в том числе агентские, подписывает необходимые документы, а также производит необходимые "
        "платежи."
    )
    draw_multiline_text(page, text, 2.5 * cm, 26.5 * cm - 0.5 * cm * 10, 17 * cm, 0.5 * cm, "Arial", 12)
    text = (
        f"2. Порядок исполнения договора"
    )
    draw_multiline_text(page, text, 2.5 * cm, 26.5 * cm - 0.5 * cm * 36, 17 * cm, 0.5 * cm, "Arial-Bold", 12)
    text = (
        f"2.1. После поступления аванса на счет Поставщика, Заказчику предоставляется доступ к торговой системе "
        f"и инструкция по ее использованию. В торговой системе Заказчик имеет возможность формировать ставки на "
        f"выбранные ТС, запрашивать переводы аукционных листов, получать дополнительную информацию и консультации "
        f"по интересующим ТС."
    )
    draw_multiline_text(page, text, 2.5 * cm, 26.5 * cm - 0.5 * cm * 37, 17 * cm, 0.5 * cm, "Arial", 12)
    text = (
        f"3. Права и обязанности сторон"
    )
    draw_multiline_text(page, text, 2.5 * cm, 26.5 * cm - 0.5 * cm * 42, 17 * cm, 0.5 * cm, "Arial-Bold", 12)
    text = (
        f"3.1. Поставщик обязуется:"
        f" ПЕРЕХОД 3.1.1. Предоставить Заказчику информационную возможность участия в легальных японских авто аукционах. "
        f" ПЕРЕХОД 3.1.2. В течение 2 суток с момента отправки ставки на заявленную Заказчиком ТС, довести до него информацию о "
        f"состоявшемся приобретении ТС."
        f" ПЕРЕХОД 3.1.3. При получении подтверждения о приобретении в пользу Заказчика указанного в договоре ТС, обеспечить "
        f"доставку данного ТС в пункт отправки (порт отправления)."
    )
    draw_multiline_text(page, text, 2.5 * cm, 26.5 * cm - 0.5 * cm * 43, 17 * cm, 0.5 * cm, "Arial", 12)

    text = (
        f" ПЕРЕХОД 3.1.4. По заявлению Заказчика организовать доставку приобретенного в его пользу автомобиля, после его "
        f"отгрузки из порта отправления и прибытия в порт назначения и по месту назначения, указанными Заказчиком."
        f" ПЕРЕХОД 3.1.5. Передать грузоотправителю приобретенное Заказчиком ТС, после оплаты им всех понесенных Поставщиком "
        f"затрат и расходов, связанных с приобретением и доставкой приобретенного Заказчиком ТС"
        f" ПЕРЕХОД 3.1.6. В течение 10-ти банковских дней вернуть сумму аванса, по желанию Заказчика, при несостоявшейся "
        f"сделке на аукционе по причине недостаточности предложенной им цены. Аванс возвращается за вычетом комиссии "
        f"банка на перевод в пользу физ.лица."
        f" ПЕРЕХОД 3.2. Заказчик обязуется:"
        f" ПЕРЕХОД 3.2.1. Предоставить Поставщику заявку (поручение) с указанием всех необходимых сведений, необходимых для "
        f"исполнения настоящего договора."
        f" ПЕРЕХОД 3.2.2. Внести Поставщику в обеспечение исполнения настоящего договора сумму аванса, а также оплатить "
        f"стоимость приобретаемого в его пользу ТС и затраты, связанные с его приобретением, а также вознаграждение "
        f"Поставщика в сроки и порядке, указанные в разделе 5 настоящего договора."
        f" ПЕРЕХОД 3.2.3. Получив от Поставщика информацию о состоявшейся или не состоявшейся сделке по приобретению заказанного "
        f"им ТС, подтвердить ее получение, направив в адрес Поставщика подтверждение по телефону, электронной почтой "
        f"или SMS сообщением."
    )
    x = 2.5 * cm
    y = 24.5 * cm
    page_width, page_height = page._pagesize
    space_left = page_height - y

    # Если текст займет больше места, чем осталось на текущей странице, добавляем новую страницу
    if page.stringWidth(text) > space_left:
        page.showPage()
        y = page_height - 2 * cm  # Новое значение y для начала новой страницы

    # Продолжаем заполнение на новой странице
    draw_multiline_text(page, text, x, y, 17 * cm, 0.5 * cm, "Arial", 12)

    text = (
        f"4. Сроки исполнения отдельных поручений Заказчика"
    )
    draw_multiline_text(page, text, 2.5 * cm, 17 * cm - 0.5 * cm * 4, 17 * cm, 0.5 * cm, "Arial-Bold", 12)
    text = (
        f"4.1. Поставщик обязуется осуществить доставку ТС, приобретенного в пользу Заказчика из порта отправки в порт "
        f"назначения (г. Владивосток) в срок, не превышающий 60 (шестидесяти) суток с момента получения от Заказчика "
        f"денежных средств в оплату аукционной стоимости ТС, стоимости доставки ТС из порта отправки в порт назначения "
        f"(г. Владивосток) и др. расходов на японской стороне."
        f" ПЕРЕХОД 4.2. Срок таможенного оформления и получения ТС с таможенного склада устанавливается Поставщиком с учетом "
        f"загруженности таможни, но не более 30 (тридцати) суток с момента поступления ТС на таможенный склад. Оплата "
        f"таможенных платежей производится согласно таможенного приходного ордера с учетом стоимости услуг по "
        f"оформлению автомобиля и входит в общую сумму договора."
        f" ПЕРЕХОД 4.3. Общий срок доставки ТС во Владивосток составляет не более 60 суток, с момента покупки ТС "
        f"на аукционе Японии."
    )
    draw_multiline_text(page, text, 2.5 * cm, 17 * cm - 0.5 * cm * 5, 17 * cm, 0.5 * cm, "Arial", 12)

    text = (
        f"5. Порядок расчетов по договору"
    )
    draw_multiline_text(page, text, 2.5 * cm, 17 * cm - 0.5 * cm * 18, 17 * cm, 0.5 * cm, "Arial-Bold", 12)

    text = (
        f"5.1. Оплата вознаграждения Поставщика и расходов на приобретение и доставку ТС,"
        f"приобретаемого в пользу Заказчика осуществляется Заказчиком в следующем порядке:"
        f" ПЕРЕХОД 5.1.1. После заключения настоящего договора, Заказчик вносит Поставщику сумму аванса в размере не менее "
        f"10 (десяти) процентов от общей предполагаемой стоимости автомобиля Заказчика, но не менее 100000 (ста тысяч) "
        f"рублей, необходимую для участия заявки Поставщика на приобретение выбранного Заказчиком ТС на аукционе."
        f" ПЕРЕХОД 5.1.2. В течение 3 (трех) банковских дней с момента уведомления Поставщиком Заказчика о состоявшемся в его "
        f"пользу аукционе, по результатам которого ТС было приобретено Поставщиком по заявке Заказчика, Заказчик "
        f"оплачивает аукционную стоимость ТС за вычетом аванса уплаченного в соответствии с п. 5.1.1, стоимость "
        f"доставки из порта отправки в Японии в порт г. Владивосток и затрат в Японии в соответствии с C&F-тарифом "
        f"компании, на основании результатов торгов и выставленного счета Поставщика."
    )
    draw_multiline_text(page, text, 2.5 * cm, 17 * cm - 0.5 * cm * 19, 17 * cm, 0.5 * cm, "Arial", 12)
    x = 2.5 * cm
    y = 24.5 * cm

    # Если текст займет больше места, чем осталось на текущей странице, добавляем новую страницу
    if page.stringWidth(text) > space_left:
        page.showPage()
        y = page_height - 2 * cm  # Новое значение y для начала новой страницы

    text = (
        f" ПЕРЕХОД 5.1.3. В течение 3 (трех) банковских дней с момента уведомления Поставщиком Заказчика об отправке "
        f"ТС в порт назначения (г. Владивосток), Заказчик оплачивает Поставщику стоимость таможенного оформления ТС, "
        f"другие расходы, а также агентское вознаграждение. Основанием для оплаты является выставленный счет "
        f"Поставщика Заказчику."
        f" ПЕРЕХОД 5.2. Форма расчетов - наличные и/или перечисление на банковский счет."
        f" ПЕРЕХОД 5.3. После выполнения Поставщиком поручения Заказчика в полном объеме, стороны производят уточненный расчет "
        f"цены договора и проводят окончательные взаиморасчеты."
    )
    # Продолжаем заполнение на новой странице
    draw_multiline_text(page, text, x, y, 17 * cm, 0.5 * cm, "Arial", 12)

    text = (
        f"6. Ответственность сторон"
    )
    draw_multiline_text(page, text, 2.5 * cm, y - 0.5 * cm * 10, 17 * cm, 0.5 * cm, "Arial-Bold", 12)

    text = (
        f"6.1. При просрочке платежей предусмотренных п.5.1.2. и 5.1.3. настоящего договора, Заказчик обязан уплатить "
        f"Поставщику до получения ТС неустойку в размере 0,33% от суммы просроченного платежа, за каждый день просрочки, "
        f"но общая сумма неустойки не может превышать 5%. ТС Заказчику не передается до уплаты им пени за просроченный платеж."
        f" ПЕРЕХОД 6.2. При отказе Заказчика от внесения платежей после покупки ТС на аукционе, настоящий договор считается "
        f"расторгнутым. Заказчик обязан возместить Поставщику прямые убытки, причиненные расторжением настоящего договора "
        f"в размере внесённого первоначального аванса (десяти) процентов от общей предполагаемой стоимости автомобиля Заказчика "
        f"(не включая стоимость отправки)"
        f" ПЕРЕХОД 6.3. В случае просрочки выполнения Поставщиком обязательств по настоящему договору, последний уплачивает "
        f"Заказчику пени в размере 0,1% от стоимости услуг Поставщика за каждый день просрочки."
        f" ПЕРЕХОД 6.4. Сторона, виновная в нанесении убытков другой стороне, возмещает их в полном объеме."
        f" ПЕРЕХОД 6.5. Меры ответственности сторон, не предусмотренные в настоящем договоре, применяются в соответствии "
        f"с нормами гражданского законодательства действующими в РФ."
        f" ПЕРЕХОД 6.6. При обнаружении в приобретенном Поставщиком ТС новых дефектов лакокрасочного покрытия кузова, "
        f"полученных в процессе доставки до Владивостока, Поставщик по требованию Заказчика:"
        f" ПЕРЕХОД a) безвозмездно устраняет существенные недостатки в приобретенном ТС;"
        f" ПЕРЕХОД б) соответственно уменьшает вознаграждение Поставщика."
        f" ПЕРЕХОД 6.7. Заказчик вправе расторгнуть настоящий Договор и потребовать от Поставщика полученной последним "
        f"оплаты стоимости ТС (или ее части), если обнаруженные существенные недостатки в поставленном ТС не были устранены "
        f"или возмещены расходы на их устранение Поставщиком в течение 30 дней с момента предъявления такого требования Заказчиком."
        f" ПЕРЕХОД 6.8. Поставщик несет ответственность за сохранность авто до момента передачи авто лично в руки или "
        f"представителям транспортной компании."
        f" ПЕРЕХОД 6.9. Поставщик не несет ответственность за возможные изменения таможенных пошлин и сроков оплаты/исполнения "
        f"таможенного оформления."
        f" ПЕРЕХОД 6.10.Поставщик не несет ответственность за наличие и сохранность вложений (колеса, детали авто, ключи, "
        f"SD карты, сервисные книжки и пр.)."
        f" ПЕРЕХОД 6.11.Поставщик не несет ответственность за новые дефекты, которые могут появиться в результате "
        f"крепления машины в процессе транспортировки (следы от строп на кузове, сколы или царапины на колесных дисках)."
    )
    draw_multiline_text(page, text, 2.5 * cm, y - 0.5 * cm * 11, 17 * cm, 0.5 * cm, "Arial", 12)

    x = 2.5 * cm
    y = 24.5 * cm
    page.showPage()
    y = page_height - 2 * cm  # Новое значение y для начала новой страницы

    text = (
        f"7. Порядок разрешения споров"
    )
    draw_multiline_text(page, text, x, y, 17 * cm, 0.5 * cm, "Arial-Bold", 12)
    text = (
        f"7.1. Любые разногласия, возникшие в процессе исполнения настоящего договора, разрешаются сторонами "
        f"путем переговоров. При не достижении соглашения, заинтересованная сторона вправе обратиться в суд в "
        f"установленном законом порядке."
        f" ПЕРЕХОД 7.2. Стороны не имеют права отказаться от исполнения настоящего договора в одностороннем порядке "
        f"(за исключением п.3.1.6., 6.2., 6.7. настоящего договора). Любые дополнения и изменения действительны, если "
        f"они совершены в письменном виде и подписаны полномочными представителями сторон."
    )
    draw_multiline_text(page, text, x, y - 0.5 * cm * 1, 17 * cm, 0.5 * cm, "Arial", 12)
    text = (
        f"8. Особые условия"
    )
    draw_multiline_text(page, text, x, y - 0.5 * cm * 9, 17 * cm, 0.5 * cm, "Arial-Bold", 12)
    text = (
        f"8.1. В случае расторжения сторонами договора, кроме п.6.7., по инициативе Заказчика, с уплаченного "
        f"им аванса удерживаются понесенные Поставщиком расходы."
        f" ПЕРЕХОД 8.2. С момента подписания настоящего договора, Заказчик при разрешении каких-либо возникающих "
        f"связанных с выполнением договора вопросов, ссылается на номер и дату настоящего договора."
        f" ПЕРЕХОД 8.3. Поставщик освобождается от ответственности за неисполнение или ненадлежащее исполнение своих "
        f"обязанностей по настоящему Договору, если докажет, что неисполнение или ненадлежащее исполнение произошло "
        f"вследствие влияния обстоятельств непреодолимой силы."
    )
    draw_multiline_text(page, text, x, y - 0.5 * cm * 10, 17 * cm, 0.5 * cm, "Arial", 12)
    text = (
        f"9. Срок действия договора и заключительные положения"
    )
    draw_multiline_text(page, text, x, y - 0.5 * cm * 20, 17 * cm, 0.5 * cm, "Arial-Bold", 12)
    text = (
        f"9.1. Договор вступает в законную силу с момента его подписания сторонами и действует до момента исполнения "
        f"обязательств сторонами в полном объеме."
        f" ПЕРЕХОД 9.2. Договор считается исполненным с момента получения автомобиля Заказчиком."
        f" ПЕРЕХОД 9.3. Настоящий договор составлен на русском языке в двух экземплярах, имеющих одинаковую юридическую "
        f"силу, по одному для каждой из сторон. Стороны намерены заключить настоящий договор посредством факсимильной "
        f"связи и/или обмена отсканированными подписанными и скрепленными печатью копиями договоров по электронной "
        f"почте, а подписи сторон на документах, переданных и полученных по факсу и электронной почте, имеют силу "
        f"собственноручных."
    )
    draw_multiline_text(page, text, x, y - 0.5 * cm * 21, 17 * cm, 0.5 * cm, "Arial", 12)
    text = (
        f"10. Форс-мажорные обстоятельства"
    )
    draw_multiline_text(page, text, x, y - 0.5 * cm * 30, 17 * cm, 0.5 * cm, "Arial-Bold", 12)
    text = (
        f"10.1. При возникновении обстоятельств, которые мешают частичному или полному выполнению обязательств по "
        f"настоящему Договору одной из сторон, а к таким обязательствам относятся: наводнения, землетрясения, другие "
        f"стихийные бедствия, а также войны, СВО, моратории, перебои в работе железнодорожного и морского транспорта, "
        f"отсутствии мест на судах при отправке из Японии, отсутствии вагонов, автовозов при отправке из Владивостока - "
        f"срок выполнения обязательств по настоящему Договору будет увеличен на период времени, в течение которого "
        f"будут действовать эти обстоятельства и последствия."
        f" ПЕРЕХОД 10.2. Если вышеуказанные обстоятельства длятся более 6 месяцев, каждая сторона (Поставщик и "
        f"Заказчик) имеет право отменить дальнейшее выполнение настоящего Договора, и в этом случае ни одна из сторон "
        f"не может требовать компенсации за причиненный ущерб. Сторона, для которой стало невозможным дальнейшее "
        f"выполнение обязательств по настоящему Договору, должна в 10-дневный срок информировать другую сторону о "
        f"начале и окончания действия обстоятельств, препятствующих дальнейшему выполнению обязательств."
    )
    draw_multiline_text(page, text, x, y - 0.5 * cm * 31, 17 * cm, 0.5 * cm, "Arial", 12)

    x = 2.5 * cm
    y = 24.5 * cm
    page.showPage()
    y = page_height - 2 * cm  # Новое значение y для начала новой страницы

    text = (
        f"11. Дополнительные условия и заключительные положения"
    )
    draw_multiline_text(page, text, x, y - 0.5 * cm * 0, 17 * cm, 0.5 * cm, "Arial-Bold", 12)
    text = (
        f"11.1. Любые изменения и дополнения к настоящему договору действительны, при условии, если они совершены в "
        f"письменной форме и подписаны сторонами или надлежаще уполномоченными, на то представителями сторон."
        f" ПЕРЕХОД 11.2. Во всем остальном, что не предусмотрено настоящим договором и Руководством Клиента, стороны "
        f"руководствуются действующим законодательством РФ."
        f" ПЕРЕХОД 11.3. Договор составлен в двух экземплярах, из которых один находится у Агента, а второй у Клиента."
        f" ПЕРЕХОД 11.4. В соответствии со ст. 9 Федерального закона от 27.07.2006 No 152-ФЗ «О персональных данных», "
        f"Клиент предоставляет Агенту право на обработку его персональных данных, а именно: ФИО, адрес, телефон, факс, "
        f"электронный адрес, дата рождения; данные об Автомобиле - VIN, модель и т.д."
    )
    draw_multiline_text(page, text, x, y - 0.5 * cm * 1, 17 * cm, 0.5 * cm, "Arial", 12)

    text = (
        f"12. Юридические адреса сторон и банковские реквизиты"
    )
    draw_multiline_text(page, text, x, y - 0.5 * cm * 0, 17 * cm, 0.5 * cm, "Arial-Bold", 12)
    probel = "                                                                                                   "
    page.drawString(2.5 * cm, y - 0.5 * cm * 13, f"Агент{probel}Клиент")
    page.drawString(2.5 * cm, 26.5 * cm - cm,
                    f"ООО «NAHODKA MOTORS»{probel}{order.id_customer.last_name_client} {order.id_customer.first_name_client} {order.id_customer.patronymic_client}")
    page.drawString(2.5 * cm, 26.5 * cm - cm, f"ИНН{probel}{order.id_customer.date_of_birth}")
    page.drawString(2.5 * cm, 26.5 * cm - cm,
                    f"ОГРНИП{probel}{order.id_customer.passport_series}№{order.id_customer.passport_number}")
    page.drawString(2.5 * cm, 26.5 * cm - cm, f"ИНН{probel}{order.id_customer.date_of_birth}")
    page.drawString(2.5 * cm, 26.5 * cm - cm, f"Р/С{probel}{order.id_customer.passport_department_name}")
    page.drawString(2.5 * cm, 26.5 * cm - cm, f"К/С{probel}{order.id_customer.date_of_birth}")
    page.drawString(2.5 * cm, 26.5 * cm - cm,
                    f"ФИЛИАЛ “ЦЕНТРАЛЬНЫЙ” БАНКА ВТБ {probel}{order.id_customer.date_of_birth}")
    page.drawString(2.5 * cm, 26.5 * cm - cm, f"(ПАО){probel}{order.id_customer.date_of_birth}")
    page.drawString(2.5 * cm, 26.5 * cm - cm, f"г. Москва, БИК 044525411{probel}{order.id_customer.date_of_birth}")
    page.drawString(2.5 * cm, 26.5 * cm - cm,
                    f"Юридический адрес:Приморский край{probel}{order.id_customer.date_of_birth}")
    page.drawString(2.5 * cm, 26.5 * cm - cm,
                    f"г. Владивосток, ул. Авроровская 19А{probel}{order.id_customer.date_of_birth}")
    page.drawString(2.5 * cm, 26.5 * cm - cm, f"Контактный тел.: 8908148193{probel}{order.id_customer.date_of_birth}")
    page.drawString(2.5 * cm, 26.5 * cm - cm, f"Эл. почта: fefucar@fefucar.ru{probel}{order.id_customer.date_of_birth}")
    page.drawString(2.5 * cm, 26.5 * cm - cm, f"Директор{probel}{order.id_customer.date_of_birth}")
    page.drawString(2.5 * cm, 26.5 * cm - cm, f"_____________/{probel}_____________/")
    page.drawString(2.5 * cm, 26.5 * cm - cm,
                    f"{order.id_worker.full_name}{probel}{order.id_customer.first_name_client} {order.id_customer.last_name_client} {order.id_customer.patronymic_client}")
    page.setFont("Arial", 10)
    page.drawString(2.5 * cm, 26.5 * cm - cm, f"подпись{probel}подпись")

    # draw_multiline_text(page, text, x, y - 0.5 * cm * 1, 17 * cm, 0.5 * cm, "Arial", 12)

    page.save()

    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f'dogovot.pdf')

# order = Order.objects.get(id_order=order_id)
